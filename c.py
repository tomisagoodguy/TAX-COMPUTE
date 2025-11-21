import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- 檔案設定 ---
FILENAME = "履保結餘款分配明細表_最終版.xlsx"

# --- 莫蘭迪商務配色 ---
COLOR_HEADER_BG = "34495E"      # 標題
COLOR_HEADER_TEXT = "FFFFFF"    # 標題白字
COLOR_CATEGORY_BG = "ECF0F1"    # 左側項目欄
COLOR_INPUT_BG = "FEF9E7"       # 輸入區 (羊皮紙色)
COLOR_TOTAL_BG = "E5E8E8"       # 自動計算區 (灰)
COLOR_HIGHLIGHT_BG = "FCF3CF"   # 結餘款 (金)
COLOR_ALERT_TEXT = "C0392B"     # 警告紅字

# --- 樣式定義 ---
border_color = "7F8C8D"
thin_border = Border(
    left=Side(style='thin', color=border_color),
    right=Side(style='thin', color=border_color),
    top=Side(style='thin', color=border_color),
    bottom=Side(style='thin', color=border_color)
)

font_header = Font(name='微軟正黑體', bold=True, color=COLOR_HEADER_TEXT, size=11)
font_category = Font(name='微軟正黑體', bold=True, color="2C3E50")
font_normal = Font(name='微軟正黑體')
font_red = Font(name='微軟正黑體', bold=True, color=COLOR_ALERT_TEXT)
font_total = Font(name='微軟正黑體', bold=True)

fill_header = PatternFill(start_color=COLOR_HEADER_BG,
                          end_color=COLOR_HEADER_BG, fill_type="solid")
fill_category = PatternFill(
    start_color=COLOR_CATEGORY_BG, end_color=COLOR_CATEGORY_BG, fill_type="solid")
fill_input = PatternFill(start_color=COLOR_INPUT_BG,
                         end_color=COLOR_INPUT_BG, fill_type="solid")
fill_total = PatternFill(start_color=COLOR_TOTAL_BG,
                         end_color=COLOR_TOTAL_BG, fill_type="solid")
fill_highlight = PatternFill(
    start_color=COLOR_HIGHLIGHT_BG, end_color=COLOR_HIGHLIGHT_BG, fill_type="solid")

align_center = Alignment(
    horizontal='center', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_cell(cell, font=None, fill=None, alignment=None, number_format=None, border=None, locked=None):
    """ 統一設定儲存格樣式的輔助函式 """
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    if border:
        cell.border = border
    if locked is not None:
        cell.protection = Protection(locked=locked)


def create_distribution_sheet(wb, num_people, tab_color):
    """ 建立單一工作表 (含公式與保護設定) """
    sheet_name = f"{num_people}人分配表"
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color

    # --- 欄位計算 ---
    start_col_idx = 2  # Column B
    end_col_idx = start_col_idx + num_people - 1
    total_col_idx = end_col_idx + 1
    total_col_letter = get_column_letter(total_col_idx)
    start_col_letter = get_column_letter(start_col_idx)

    # --- 1. 總價金區域 (B1) ---
    ws['A1'] = "價金總額"
    ws['B1'] = 18300000

    style_cell(ws['A1'], font=font_header, fill=fill_header,
               alignment=align_center, border=thin_border)
    # B1 設定為可輸入 (locked=False)
    style_cell(ws['B1'], font=Font(name='微軟正黑體', bold=True, size=12), fill=fill_input,
               number_format='#,##0', alignment=align_right, border=thin_border, locked=False)

    # 根據人數決定合併寬度
    if num_people > 4:
        ws.merge_cells(f'B1:{get_column_letter(start_col_idx+2)}1')
    else:
        ws.merge_cells('B1:C1')

    ws.column_dimensions['A'].width = 24
    ws.row_dimensions[1].height = 25

    # --- 2. 標題列 (Row 3) ---
    ws['A3'] = "項目"
    style_cell(ws['A3'], font=font_header, fill=fill_header,
               alignment=align_center, border=thin_border)
    ws.row_dimensions[3].height = 32

    for i in range(num_people):
        col_letter = get_column_letter(start_col_idx + i)
        ws[f'{col_letter}3'] = f'人員{chr(65+i)}'
        style_cell(ws[f'{col_letter}3'], font=font_header,
                   fill=fill_header, alignment=align_center, border=thin_border)
        ws.column_dimensions[col_letter].width = 14

    ws[f'{total_col_letter}3'] = "合計"
    style_cell(ws[f'{total_col_letter}3'], font=font_header,
               fill=fill_header, alignment=align_center, border=thin_border)
    ws.column_dimensions[total_col_letter].width = 16

    # --- 3. 資料結構設定 ---
    # 格式：(項目名稱, 數值格式, 是否可輸入, 是否為分隔線)
    data_rows_config = [
        ('持分', '# ?/???', True, False),      # [重要] 設定為分數格式，支援 1/3 顯示
        ('分配金額', '#,##0', False, False),
        (None, None, False, False),
        ('--- 支出項目 ---', None, False, True),
        ('仲介費', '#,##0', True, False),
        ('土增稅', '#,##0', True, False),
        ('繳稅匯費', '#,##0', True, False),
        ('地價稅', '#,##0', True, False),
        ('房屋稅', '#,##0', True, False),
        ('稅費分算', '#,##0', True, False),
        ('二胎', '#,##0', True, False),
        ('水電費', '#,##0', True, False),
        ('代書費', '#,##0', True, False),
        ('銀行貸款', '#,##0', True, False),
        ('其他(可自填)1', '#,##0', True, False),
        ('其他(可自填)2', '#,##0', True, False),
        ('其他(可自填)3', '#,##0', True, False),
        (None, None, False, False),
        ('支出項目合計', '#,##0', False, False),
        (None, None, False, False),
        ('履保結餘款', '#,##0', False, False),
    ]

    current_row = 4
    rows_map = {}
    expense_start_row = -1
    expense_end_row = -1

    for item, num_format, is_input, is_separator in data_rows_config:
        if item is None:
            current_row += 1
            continue

        rows_map[item] = current_row
        if item == '仲介費':
            expense_start_row = current_row
        if '其他' in item:
            expense_end_row = current_row

        # A欄 (項目名稱)
        ws[f'A{current_row}'] = item

        if is_separator:
            # 分隔線樣式
            style_cell(ws[f'A{current_row}'], font=font_category,
                       fill=fill_category, alignment=align_left)
            ws.merge_cells(f'A{current_row}:{total_col_letter}{current_row}')
            style_cell(ws[f'A{current_row}'], font=Font(name='微軟正黑體', bold=True, color="2C3E50", italic=True),
                       fill=fill_category, alignment=align_center, border=thin_border)
        else:
            # 一般列樣式
            is_balance = (item == '履保結餘款')
            f_style = font_red if is_balance else font_category
            bg_style = fill_highlight if is_balance else fill_category
            style_cell(ws[f'A{current_row}'], font=f_style,
                       fill=bg_style, alignment=align_left, border=thin_border)

            # 數值欄位迴圈 (B欄 ~ 合計欄)
            for i in range(start_col_idx, total_col_idx + 1):
                col = get_column_letter(i)
                cell = ws[f'{col}{current_row}']
                is_total_col = (i == total_col_idx)

                if is_total_col:
                    # 合計欄 (鎖定)
                    bg = fill_highlight if is_balance else fill_total
                    ft = font_red if is_balance else font_total
                    # 持分合計改為百分比顯示
                    fmt = '0.00%' if item == '持分' else num_format
                    style_cell(cell, font=ft, fill=bg, number_format=fmt,
                               alignment=align_right, border=thin_border, locked=True)
                else:
                    # 個人欄
                    if is_input:
                        # [重要] 輸入欄位解除鎖定 (locked=False)
                        style_cell(cell, font=font_normal, fill=fill_input, number_format=num_format,
                                   alignment=align_right, border=thin_border, locked=False)
                    else:
                        # 計算欄位鎖定
                        bg = fill_highlight if is_balance else fill_total
                        ft = font_red if is_balance else font_normal
                        style_cell(cell, font=ft, fill=bg, number_format=num_format,
                                   alignment=align_right, border=thin_border, locked=True)
        current_row += 1

    # --- 4. 寫入公式 (解決尾差的核心邏輯) ---
    row_share = rows_map['持分']
    row_alloc = rows_map['分配金額']
    row_exp_total = rows_map['支出項目合計']
    row_balance = rows_map['履保結餘款']
    total_price_ref = '$B$1'

    for i in range(num_people):
        col_idx = start_col_idx + i
        col = get_column_letter(col_idx)
        share_cell = f'{col}{row_share}'

        # [平帳邏輯]
        if i == 0:
            # 第一人：正常計算四捨五入
            # =IF(B4<>"", ROUND(B4*$B$1, 0), 0)
            formula_alloc = f'=IF({share_cell}<>"", ROUND({share_cell}*{total_price_ref}, 0), 0)'
        else:
            # 後續人：(累計持分 * 總價) - 已分配總額
            # 這確保了 最後一人金額 = 總價 - 前面所有人的總和
            cum_share_range = f'{start_col_letter}{row_share}:{col}{row_share}'
            prev_alloc_range = f'{start_col_letter}{row_alloc}:{get_column_letter(col_idx-1)}{row_alloc}'

            formula_alloc = (f'=IF({share_cell}<>"", '
                             f'ROUND(SUM({cum_share_range})*{total_price_ref}, 0) - SUM({prev_alloc_range}), '
                             f'0)')

        ws[f'{col}{row_alloc}'] = formula_alloc

        # 支出與結餘公式
        ws[f'{col}{row_exp_total}'] = f'=IF({share_cell}<>"", SUM({col}{expense_start_row}:{col}{expense_end_row}), 0)'
        ws[f'{col}{row_balance}'] = f'=IF({share_cell}<>"", {col}{row_alloc}-{col}{row_exp_total}, 0)'

    # 右側合計欄公式
    calc_rows = [row_share, row_alloc, row_exp_total, row_balance] + \
        list(range(expense_start_row, expense_end_row + 1))
    for r in calc_rows:
        ws[f'{total_col_letter}{r}'] = f'=SUM({start_col_letter}{r}:{get_column_letter(end_col_idx)}{r})'

    # --- 5. 簽名區與美化 ---
    sig_row = current_row + 2
    ws.merge_cells(f'A{sig_row-1}:{total_col_letter}{sig_row-1}')
    style_cell(ws[f'A{sig_row-1}'],
               border=Border(top=Side(style='medium', color="7F8C8D")))

    ws[f'A{sig_row}'] = "上述金額經各共有人確認無誤。"
    style_cell(ws[f'A{sig_row}'], font=Font(
        name='微軟正黑體', bold=True, size=12), alignment=align_left)

    ws[f'{get_column_letter(total_col_idx-2)}{sig_row}'] = "簽名："
    style_cell(ws[f'{get_column_letter(total_col_idx-2)}{sig_row}'],
               font=Font(name='微軟正黑體', bold=True, size=12), alignment=align_right)

    ws[f'A{sig_row + 2}'] = "中華民國        年        月        日"
    ws.merge_cells(f'A{sig_row + 2}:{total_col_letter}{sig_row + 2}')
    style_cell(ws[f'A{sig_row + 2}'], font=font_normal, alignment=align_right)

    # --- 6. 啟用保護 (關鍵修改：開放格式權限) ---
    ws.protection.sheet = True
    ws.protection.password = '5168'

    # [關鍵] 允許使用者在保護狀態下修改格式、調整欄寬列高
    ws.protection.formatCells = True
    ws.protection.formatColumns = True
    ws.protection.formatRows = True

    ws.protection.enable()
    print(f"  - 工作表 '{sheet_name}' 建立完成 (平帳公式 + 允許格式修改)")


# --- 主程式 ---
wb = openpyxl.Workbook()
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# 生成 4, 6, 8, 10 人表
configs = [(4, "1ABC9C"), (6, "2ECC71"), (8, "F39C12"), (10, "E74C3C")]

for num, color in configs:
    create_distribution_sheet(wb, num, color)

try:
    wb.save(FILENAME)
    print(f"\n✨ 成功建立檔案： '{FILENAME}'")
    print("💡 新功能：")
    print("   1. 在持分欄位輸入 '1/3'，會自動顯示為分數並精確計算。")
    print("   2. 右鍵 '儲存格格式' 功能已解鎖，您可以自由調整字體或顏色。")
except PermissionError:
    print(f"❌ 失敗：請先關閉 '{FILENAME}' 再執行。")
