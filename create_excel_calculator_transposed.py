import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- Configuration ---
FILENAME = "履保結餘款分配明細表.xlsx"
SHEET_NAME = "結餘款分配明細"

# --- Create Workbook and Sheet ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = SHEET_NAME

# --- Define Styles ---
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
input_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
category_font = Font(bold=True)
category_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
total_font = Font(bold=True)
total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
red_font = Font(bold=True, color="FF0000")
center_wrap_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_wrap_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

# --- Helper to apply styles ---
def style_cell(cell, font=None, fill=None, alignment=None, number_format=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if number_format: cell.number_format = number_format
    if border: cell.border = border

# --- Main Price Input ---
ws['A1'] = "價金總額"
ws['B1'] = 18300000
style_cell(ws['A1'], font=header_font, fill=header_fill, alignment=center_wrap_alignment)
style_cell(ws['B1'], fill=input_fill, number_format='#,##0', alignment=right_wrap_alignment)
ws.column_dimensions['A'].width = 20
ws.row_dimensions[1].height = 20

# --- Headers for People & Total ---
ws['A3'] = "項目"
style_cell(ws['A3'], font=header_font, fill=header_fill, alignment=center_wrap_alignment)
ws.row_dimensions[3].height = 30

people_headers = ['人員A', '人員B', '人員C', '人員D', '人員E', '人員F', '合計']
for i, person in enumerate(people_headers, start=2):
    col_letter = get_column_letter(i)
    ws[f'{col_letter}3'] = person
    style_cell(ws[f'{col_letter}3'], font=header_font, fill=header_fill, alignment=center_wrap_alignment)
    ws.column_dimensions[col_letter].width = 15

# --- Data Rows ---
data_rows = [
    ('持分', '# ?/?'),
    ('分配金額', '#,##0'),
    (None, None),
    ('--- 支出項目 ---', None),
    ('仲介費', '#,##0'),
    ('土增稅', '#,##0'),
    ('繳稅匯費', '#,##0'),
    ('地價稅', '#,##0'),
    ('房屋稅', '#,##0'),
    ('稅費分算', '#,##0'),
    ('二胎', '#,##0'),
    ('水電費', '#,##0'),
    ('代書費', '#,##0'),
    ('銀行貸款', '#,##0'),
    (None, None),
    ('支出項目合計', '#,##0'),
    (None, None),
    ('履保結餘款', '#,##0'),
]

current_row = 4
for item, num_format in data_rows:
    if item is None:
        current_row += 1
        continue
    
    ws[f'A{current_row}'] = item
    font_style = red_font if item == '履保結餘款' else category_font
    if '---' in item:
        style_cell(ws[f'A{current_row}'], font=font_style, alignment=left_wrap_alignment)
    else:
        style_cell(ws[f'A{current_row}'], font=font_style, fill=category_fill, alignment=left_wrap_alignment)

    for i in range(2, 8):
        col_letter = get_column_letter(i)
        cell = ws[f'{col_letter}{current_row}']
        is_input_cell = item in ['持分', '仲介費', '土增稅', '繳稅匯費', '地價稅', '房屋稅', '稅費分算', '二胎', '水電費', '代書費', '銀行貸款']
        font_override = red_font if item == '履保結餘款' else None
        if is_input_cell:
            style_cell(cell, fill=input_fill, number_format=num_format, alignment=right_wrap_alignment)
        else:
            style_cell(cell, font=font_override, number_format=num_format, alignment=right_wrap_alignment)
    current_row += 1

# --- Add initial data ---
ws['B4'] = 0.5
ws['C4'] = 0.5

# --- Add Formulas ---
for i in range(2, 8):
    col_letter = get_column_letter(i)
    ws[f'{col_letter}5'].value = f'=IF({col_letter}4<>"", {col_letter}4*$B$1, "")'
    ws[f'{col_letter}19'].value = f'=IF({col_letter}4<>"", SUM({col_letter}8:{col_letter}17), "")'
    ws[f'{col_letter}21'].value = f'=IF({col_letter}4<>"", {col_letter}5-{col_letter}19, "")'

# --- Add Formulas for "Total" Column ---
total_col_letter = 'H'
row_indices_to_sum = [5, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 21]
for row_idx in row_indices_to_sum:
    formula = f'=SUM(B{row_idx}:G{row_idx})'
    cell = ws[f'{total_col_letter}{row_idx}']
    cell.value = formula
    font_override = red_font if row_idx == 21 else total_font
    style_cell(cell, font=font_override, fill=total_fill, number_format='#,##0', alignment=right_wrap_alignment)

ws['H4'].value = f'=SUM(B4:G4)'
style_cell(ws['H4'], font=total_font, fill=total_fill, number_format='0.00%', alignment=right_wrap_alignment)

# --- Data Validation ---
dv = DataValidation(type="custom", formula1="=ROUND($H$4,10)=1")
dv.errorTitle = "持分加總錯誤"
dv.error = "所有持分的總和必須等於1 (100%)。"
dv.allow_blank = True
ws.add_data_validation(dv)
dv.add('B4:G4')

# --- Apply Borders ---
max_calc_row = 21
for row in ws.iter_rows(min_row=1, max_row=max_calc_row, min_col=1, max_col=8):
    for cell in row:
        if ws.cell(row=cell.row, column=1).value is not None or cell.row in [1, 3]:
             style_cell(cell, border=thin_border)

# --- Add Signature Block ---
signature_row_start = max_calc_row + 3
ws[f'A{signature_row_start}'] = "上述金額經各共有人確認無誤。       簽名:"
ws.merge_cells(f'A{signature_row_start}:H{signature_row_start}')
style_cell(ws[f'A{signature_row_start}'], alignment=Alignment(vertical='center'))

ws[f'A{signature_row_start + 2}'] = "中華民國        年        月        日"
ws.merge_cells(f'A{signature_row_start + 2}:H{signature_row_start + 2}')
style_cell(ws[f'A{signature_row_start + 2}'], alignment=Alignment(horizontal='center', vertical='center'))


# --- Save the workbook ---
try:
    wb.save(FILENAME)
    print(f"成功建立最終版檔案： '{FILENAME}'")
except PermissionError:
    print(f"寫入檔案錯誤： '{FILENAME}' 已開啟，請關閉後再執行一次。")
except Exception as e:
    print(f"寫入檔案時發生錯誤： {e}")
