import openpyxl
from openpyxl.styles import Font, PatternFill

# --- Configuration ---
FILENAME = "多人交屋價金分算_自動計算版.xlsx"
SHEET_NAME = "結餘款分配明細"

# --- Create Workbook and Sheet ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = SHEET_NAME

# --- Define Styles ---
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
input_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # Light yellow for inputs

# --- Helper to apply styles ---
def style_cell(cell, font=None, fill=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill

# --- First Section: Main Inputs ---
ws['A1'] = "價金總額"
ws['B1'] = 18300000
style_cell(ws['A1'], font=header_font, fill=header_fill)
style_cell(ws['B1'], fill=input_fill)
ws.column_dimensions['A'].width = 15
ws['B1'].number_format = '#,##0'


ws['A3'] = "姓名"
ws['B3'] = "持分"
style_cell(ws['A3'], font=header_font, fill=header_fill)
style_cell(ws['B3'], font=header_font, fill=header_fill)

# --- Person Data (Inputs) ---
persons = {
    'A4': ('人員A', 0.25),
    'A5': ('人員B', 0.25),
    'A6': ('人員C', 0.50),
}
for row_num in range(4, 10): # Rows 4 to 9 for 6 people
    cell_ref = f'A{row_num}'
    if cell_ref in persons:
        name, share = persons[cell_ref]
        ws[f'A{row_num}'] = name
        ws[f'B{row_num}'] = share
    
    style_cell(ws[f'A{row_num}'], fill=input_fill)
    style_cell(ws[f'B{row_num}'], fill=input_fill)
    ws[f'B{row_num}'].number_format = '0.00%'


# --- Second Section: Calculation Table Headers ---
headers = [
    ('C3', "分配金額", 15),
    ('D3', "仲介費", 12),
    ('E3', "土增稅", 12),
    ('F3', "繳稅匯費", 12),
    ('G3', "地價稅", 12),
    ('H3', "房屋稅", 12),
    ('I3', "二胎", 12),
    ('J3', "水電費", 12),
    ('K3', "代書費", 12),
    ('L3', "銀行貸款", 12),
    ('M3', "支出項目合計", 15),
    ('N3', "履保結餘款", 15),
]

for cell_ref, text, width in headers:
    ws[cell_ref] = text
    style_cell(ws[cell_ref], font=header_font, fill=header_fill)
    ws.column_dimensions[ws[cell_ref].column_letter].width = width

# --- Third Section: Formulas and input cell styles ---
for i in range(4, 10): # Rows 4 to 9 for 6 people
    # Formula for 分配金額
    ws[f'C{i}'] = f'=IF(B{i}<>"", B{i}*$B$1, "")'
    ws[f'C{i}'].number_format = '#,##0'

    # Formula for 支出項目合計
    ws[f'M{i}'] = f'=IF(C{i}<>"", SUM(D{i}:L{i}), "")'
    ws[f'M{i}'].number_format = '#,##0'

    # Formula for 履保結餘款
    ws[f'N{i}'] = f'=IF(C{i}<>"", C{i}-M{i}, "")'
    ws[f'N{i}'].number_format = '#,##0'
    
    # Style the input expense cells (D to L)
    for col_letter in "DEFGHIJKL":
        style_cell(ws[f'{col_letter}{i}'], fill=input_fill)
        ws[f'{col_letter}{i}'].number_format = '#,##0'


# --- Save the workbook ---
try:
    wb.save(FILENAME)
    print(f"成功建立Excel檔案： '{FILENAME}'")
except PermissionError:
    print(f"寫入檔案錯誤： '{FILENAME}' 已開啟，請關閉後再執行一次。")
except Exception as e:
    print(f"寫入檔案時發生錯誤： {e}")
